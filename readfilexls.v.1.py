#!/usr/bin/python
# -*- coding: utf-8 -*-
import os,re
import configparser
from openpyxl import load_workbook
FILE="Inventario.xlsx"

def ListColumns ( num ):
    max = len(range(ord('A'), ord('Z'))) + 1
    init = ord('A')
    maxChar = 625
    Elements = []
    if num > maxChar:
        print ("error {0} exceeds {1}").format(num, maxChar)
        exit(0)
    if int(num/max) > 0:
        for i in range(1, int(num / max) * max + 1):
            if i / max == 0 or i == max:
                value = chr(i - 1 + init)
            else:
                if i == max * (i / max):
                    value = chr(((i - 1) / max) - 1 + init) + chr(max - 1 + init)
                else:
                    value = chr((i / max) - 1 + init) + chr(i - (max * (i / max)) - 1 + init)
            print ("{0:3} {1:3}").format(str(i),value)
            Elements.append(value)
        for i in range(num - ord + 1, num + 1):
            value = chr((i / max) - 1 + init) + chr(i - (max * (i / max)) - 1 + init)
            print ("{0:3} {1:3}").format(str(i), value)
            Elements.append(value)
    else:
        for i in range(1, num):
            value = value = chr(i - 1 + init)
            # print ("{0:3} {1:3}").format(str(i),value)
            Elements.append(value)
    return Elements

def read_page ( sheet ):
    print ("\tTitle: {0}").format(sheet.title)
    rows=int(sheet.max_row)
    columns=int(sheet.max_column)
    print ("\tcontents: {0:06}x{1:06}").format(rows,columns)
    for r in range(1, rows+1):
        for c in (ListColumns(columns+1)):
             value=("{0}{1}").format(c,str(r))
             content= sheet[value].value
             if content :
                print content,
        print


wb = load_workbook(filename = FILE)
sheet_ranges = wb.get_sheet_names()
for page in sheet_ranges:
    print ("Page: {0}").format(page)
    read_page( wb.get_sheet_by_name(page))
    exit(0)

